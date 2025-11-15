"""Utilities for generating demo data for quick test builds."""
from __future__ import annotations

import asyncio
import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook

from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.core.security import ROLE_PERMISSIONS
from app.database import Base
from app.models.brigade import Brigade, BrigadeDailyScore
from app.models.checklist import (
    CheckInstance,
    CheckStatus,
    ChecklistTemplate,
    ChecklistTemplateVersion,
    TemplateStatus,
)
from app.models.report import Report, ReportFormatXLSX, ReportStatus
from app.models.reporting import (
    PeriodSummaryGranularity,
    ReportGenerationEvent,
    ReportGenerationEventType,
    ReportGenerationStatus,
    ReportPeriodSummary,
    RemarkEntry,
    RemarkSeverity,
)
from app.models.schedule import Schedule
from app.models.user import Role, User
from app.models.webhook import WebhookEvent, WebhookSubscription
from app.services.bootstrap_service import (
    DEFAULT_ADMIN_EMAIL,
    DEFAULT_ADMIN_NAME,
    DEFAULT_ADMIN_PASSWORD,
    DEFAULT_ROLE_DESCRIPTIONS,
    ensure_default_admin,
    ensure_roles,
)
from app.services.auth_service import AuthService
from app.services.storage_service import storage_service


@dataclass
class DemoDataResult:
    """Collection of counters for generated demo data."""

    users_created: int = 0
    brigades_created: int = 0
    templates_created: int = 0
    checks_created: int = 0
    reports_created: int = 0
    scores_created: int = 0

    def as_payload(self, locale: str = "en") -> Dict[str, Any]:
        """Convert counters to response payload."""
        from app.localization.helpers import get_translation
        
        total = (
            self.users_created
            + self.brigades_created
            + self.templates_created
            + self.checks_created
            + self.reports_created
            + self.scores_created
        )
        status = "skipped" if total == 0 else "created"
        detail = (
            get_translation("demo.data_already_available", locale)
            if status == "skipped"
            else get_translation("demo.data_created", locale)
        )
        return {
            "status": status,
            "detail": detail,
            "created_users": self.users_created,
            "created_brigades": self.brigades_created,
            "created_templates": self.templates_created,
            "created_checks": self.checks_created,
            "created_reports": self.reports_created,
            "created_scores": self.scores_created,
            "already_populated": status == "skipped",
            "external_base_url": settings.EXTERNAL_IP.rstrip("/"),
        }


RESET_IGNORE_TABLES = {"alembic_version"}

DEMO_USERS = [
    {
        "email": "demo.inspector@example.com",
        "full_name": "Демо Инспектор",
        "password": "demo123!",
        "role": "inspector",
    },
    {
        "email": "demo.inspector2@example.com",
        "full_name": "Демо Инспектор 2",
        "password": "demo123!",
        "role": "inspector",
    },
    {
        "email": "demo.lead@example.com",
        "full_name": "Демо Руководитель",
        "password": "demo123!",
        "role": "crew_leader",
    },
    {
        "email": "demo.lead2@example.com",
        "full_name": "Демо Руководитель 2",
        "password": "demo123!",
        "role": "crew_leader",
    },
    {
        "email": "demo.viewer@example.com",
        "full_name": "Демо Наблюдатель",
        "password": "demo123!",
        "role": "viewer",
    },
    {
        "email": "demo.manager@example.com",
        "full_name": "Демо Менеджер",
        "password": "demo123!",
        "role": "admin",
    },
]

DEMO_BRIGADES = [
        {
            "name": "Demo Brigade Alpha",
            "description": "Команда по монтажу и пуско-наладке",
            "leader": "demo.lead@example.com",
            "members": [
                "demo.lead@example.com",
                "demo.inspector@example.com",
            ],
            "profile": {"project": "Project Phoenix", "shift": "A"},
        },
        {
            "name": "Demo Brigade Beta",
            "description": "Бригада по контролю качества",
            "leader": "demo.inspector@example.com",
            "members": [
                "demo.inspector@example.com",
                "demo.viewer@example.com",
            ],
            "profile": {"project": "Project Atlas", "shift": "B"},
        },
        {
            "name": "Demo Brigade Gamma",
            "description": "Бригада по техническому обслуживанию",
            "leader": "demo.lead2@example.com",
            "members": [
                "demo.lead2@example.com",
                "demo.inspector2@example.com",
            ],
            "profile": {"project": "Project Titan", "shift": "C"},
        },
        {
            "name": "Demo Brigade Delta",
            "description": "Бригада по упаковке и логистике",
            "leader": "demo.inspector2@example.com",
            "members": [
                "demo.inspector2@example.com",
                "demo.inspector@example.com",
            ],
            "profile": {"project": "Project Mercury", "shift": "A"},
        },
    ]

DEMO_TEMPLATES = [
    {
        "name": "Demo Template: Safety Walk",
        "description": "Контрольный лист для оценки безопасности на объекте.",
        "schema": {
            "sections": [
                {
                    "id": "safety",
                    "title": "Безопасность",
                    "questions": [
                        {
                            "id": "safety-gear",
                            "text": "Все сотрудники используют СИЗ?",
                            "type": "boolean",
                            "required": True,
                        },
                        {
                            "id": "hazards",
                            "text": "Есть ли выявленные опасности?",
                            "type": "text",
                            "required": False,
                            "meta": {"placeholder": "Опишите выявленные риски"},
                        },
                    ],
                },
                {
                    "id": "environment",
                    "title": "Рабочая среда",
                    "questions": [
                        {
                            "id": "cleanliness",
                            "text": "Рабочие зоны содержатся в чистоте?",
                            "type": "boolean",
                            "required": True,
                        },
                        {
                            "id": "equipment",
                            "text": "Оборудование обслужено и исправно?",
                            "type": "choice",
                            "required": True,
                            "meta": {
                                "options": [
                                    {"value": "ok", "label": "Да"},
                                    {"value": "maintenance", "label": "Требует обслуживания"},
                                    {"value": "faulty", "label": "Неисправно"},
                                ]
                            },
                        },
                    ],
                },
            ]
        },
    },
    {
        "name": "Demo Template: Final Inspection",
        "description": "Чек-лист для приемки готового проекта.",
        "schema": {
            "sections": [
                {
                    "id": "visual",
                    "title": "Визуальный осмотр",
                    "questions": [
                        {
                            "id": "finish-quality",
                            "text": "Качество отделки соответствует стандартам?",
                            "type": "boolean",
                            "required": True,
                        },
                        {
                            "id": "photos",
                            "text": "Приложены фото итогового состояния?",
                            "type": "attachment",
                            "required": False,
                        },
                    ],
                },
                {
                    "id": "documentation",
                    "title": "Документация",
                    "questions": [
                        {
                            "id": "reports",
                            "text": "Сформированы итоговые отчеты?",
                            "type": "boolean",
                            "required": True,
                        },
                        {
                            "id": "notes",
                            "text": "Комментарии",
                            "type": "text",
                            "required": False,
                            "meta": {"multiline": True},
                        },
                    ],
                },
            ]
        },
    },
    {
        "name": "Demo Template: Workspace Readiness",
        "description": "Состояние рабочего пространства укладчика-упаковщика и смежных ролей.",
        "schema": {
            "sections": [
                {
                    "id": "workspace-readiness",
                    "title": "Рабочая зона",
                    "questions": [
                        {
                            "id": "floor-cleanliness",
                            "text": "На полу отсутствует пыль и грязь, нет россыпей сырья или красителя.",
                            "type": "boolean",
                            "required": True,
                            "meta": {"critical": True, "requires_ok": True, "points": 2},
                        },
                        {
                            "id": "personal-items-removed",
                            "text": "Личные вещи отсутствуют в рабочей зоне. Работник использует головной убор и перчатки, волосы убраны.",
                            "type": "boolean",
                            "required": True,
                            "meta": {"critical": True, "requires_ok": True, "points": 2},
                        },
                        {
                            "id": "no-foreign-objects",
                            "text": "На оборудовании и рядом нет посторонних предметов, не относящихся к работе.",
                            "type": "boolean",
                            "required": True,
                            "meta": {"points": 1.5},
                        },
                        {
                            "id": "cleaning-tools-ready",
                            "text": "Уборочный инвентарь имеется и аккуратно размещен в установленном месте.",
                            "type": "boolean",
                            "required": True,
                            "meta": {"points": 1},
                        },
                        {
                            "id": "materials-zoned",
                            "text": "Тара, упаковка, поддоны и продукция находятся в разметке идеального рабочего места.",
                            "type": "boolean",
                            "required": True,
                            "meta": {"points": 1.5},
                        },
                        {
                            "id": "passages-clear",
                            "text": "Проходы вокруг оборудования свободны от загромождений.",
                            "type": "boolean",
                            "required": True,
                            "meta": {"critical": True, "requires_ok": True, "points": 2},
                        },
                    ],
                }
            ]
        },
    },
]

DEMO_CHECKS = [
    {
        "project_id": "DEMO-PROJECT-001",
        "template": "Demo Template: Safety Walk",
        "brigade": "Demo Brigade Alpha",
        "inspector": "demo.inspector@example.com",
        "status": CheckStatus.COMPLETED,
        "started_offset": {"days": 5, "hours": 3},
        "duration_hours": 2,
        "answers": {
            "safety-gear": True,
            "hazards": "Выявлены мелкие риски, устранены на месте.",
            "cleanliness": True,
            "equipment": "ok",
        },
        "report_formats": ["xlsx"],
    },
    {
        "project_id": "DEMO-PROJECT-002",
        "template": "Demo Template: Safety Walk",
        "brigade": "Demo Brigade Beta",
        "inspector": "demo.inspector@example.com",
        "status": CheckStatus.IN_PROGRESS,
        "started_offset": {"days": 2, "hours": 6},
        "duration_hours": None,
        "answers": {
            "safety-gear": True,
            "hazards": "",
            "cleanliness": False,
            "equipment": "maintenance",
        },
        "report_formats": [],
    },
    {
        "project_id": "DEMO-PROJECT-003",
        "template": "Demo Template: Final Inspection",
        "brigade": "Demo Brigade Alpha",
        "inspector": "demo.lead@example.com",
        "status": CheckStatus.COMPLETED,
        "started_offset": {"days": 1, "hours": 4},
        "duration_hours": 3,
        "answers": {
            "finish-quality": True,
            "photos": ["minio://quality-control/demo/final/alpha.jpg"],
            "reports": True,
            "notes": "Проект принят без замечаний.",
        },
        "report_formats": ["xlsx"],
    },
    {
        "project_id": "DEMO-PROJECT-004",
        "template": "Demo Template: Workspace Readiness",
        "brigade": "Demo Brigade Beta",
        "inspector": "demo.inspector@example.com",
        "status": CheckStatus.COMPLETED,
        "started_offset": {"days": 3, "hours": 5},
        "duration_hours": 1.5,
        "answers": {
            "floor-cleanliness": True,
            "personal-items-removed": True,
            "no-foreign-objects": False,
            "cleaning-tools-ready": True,
            "materials-zoned": True,
            "passages-clear": False,
        },
        "report_formats": ["xlsx"],
    },
    {
        "project_id": "DEMO-PROJECT-005",
        "template": "Demo Template: Workspace Readiness",
        "brigade": "Demo Brigade Alpha",
        "inspector": "demo.inspector@example.com",
        "status": CheckStatus.IN_PROGRESS,
        "planned_hours_ahead": 36,
        "answers": {},
        "report_formats": [],
    },
    {
        "project_id": "DEMO-PROJECT-006",
        "template": "Demo Template: Safety Walk",
        "brigade": "Demo Brigade Gamma",
        "inspector": "demo.inspector2@example.com",
        "status": CheckStatus.COMPLETED,
        "started_offset": {"days": 7, "hours": 2},
        "duration_hours": 1.5,
        "answers": {
            "safety-gear": True,
            "hazards": "Все в порядке",
            "cleanliness": True,
            "equipment": "ok",
        },
        "report_formats": ["xlsx"],
    },
    {
        "project_id": "DEMO-PROJECT-007",
        "template": "Demo Template: Final Inspection",
        "brigade": "Demo Brigade Delta",
        "inspector": "demo.lead2@example.com",
        "status": CheckStatus.COMPLETED,
        "started_offset": {"days": 4, "hours": 8},
        "duration_hours": 2.5,
        "answers": {
            "finish-quality": False,
            "photos": ["minio://quality-control/demo/final/delta.jpg"],
            "reports": True,
            "notes": "Обнаружены незначительные дефекты, требуют доработки.",
        },
        "report_formats": ["xlsx"],
    },
    {
        "project_id": "DEMO-PROJECT-008",
        "template": "Demo Template: Workspace Readiness",
        "brigade": "Demo Brigade Beta",
        "inspector": "demo.inspector@example.com",
        "status": CheckStatus.CANCELLED,
        "started_offset": {"days": 6, "hours": 1},
        "duration_hours": None,
        "answers": {},
        "report_formats": [],
    },
    {
        "project_id": "DEMO-PROJECT-009",
        "template": "Demo Template: Safety Walk",
        "brigade": "Demo Brigade Gamma",
        "inspector": "demo.inspector2@example.com",
        "status": CheckStatus.IN_PROGRESS,
        "planned_hours_ahead": 24,
        "answers": {},
        "report_formats": [],
    },
    {
        "project_id": "DEMO-PROJECT-010",
        "template": "Demo Template: Final Inspection",
        "brigade": "Demo Brigade Delta",
        "inspector": "demo.lead@example.com",
        "status": CheckStatus.COMPLETED,
        "started_offset": {"days": 10, "hours": 5},
        "duration_hours": 3,
        "answers": {
            "finish-quality": True,
            "photos": ["minio://quality-control/demo/final/delta2.jpg"],
            "reports": True,
            "notes": "Отличное качество выполнения работ.",
        },
        "report_formats": ["xlsx"],
    },
]


async def _get_or_create_users(
    db: AsyncSession,
    *,
    role_map: Dict[str, Role],
) -> Tuple[Dict[str, User], int]:
    """Create demo users if needed and return user map and created count."""
    user_map: Dict[str, User] = {}
    created_count = 0

    for payload in DEMO_USERS:
        result = await db.execute(
            select(User).where(User.email == payload["email"])
        )
        user_obj = result.scalar_one_or_none()
        if user_obj:
            user_map[payload["email"]] = user_obj
            continue

        user_obj = User(
            email=payload["email"],
            password_hash=AuthService.hash_password(payload["password"]),
            full_name=payload["full_name"],
            is_active=True,
        )
        role = role_map.get(payload["role"])
        if role:
            user_obj.roles = [role]
        db.add(user_obj)
        await db.flush()
        user_map[payload["email"]] = user_obj
        created_count += 1

    if created_count:
        await db.commit()

    return user_map, created_count


async def _create_brigades(
    db: AsyncSession,
    *,
    user_map: Dict[str, User],
) -> Tuple[Dict[str, Brigade], int, int]:
    """Create demo brigades when missing and return map plus counters."""
    brigade_map: Dict[str, Brigade] = {}
    brigades_created = 0
    scores_created = 0

    for payload in DEMO_BRIGADES:
        result = await db.execute(
            select(Brigade).where(Brigade.name == payload["name"])
        )
        brigade = result.scalar_one_or_none()
        if brigade:
            brigade_map[payload["name"]] = brigade
            continue

        leader = user_map.get(payload["leader"])
        members = [
            user_map[email]
            for email in payload["members"]
            if email in user_map
        ]

        brigade = Brigade(
            name=payload["name"],
            description=payload["description"],
            leader_id=leader.id if leader else None,
            is_active=True,
            profile=payload.get("profile") or {},
        )
        brigade.members = members
        db.add(brigade)
        await db.flush()

        brigade_map[payload["name"]] = brigade
        brigades_created += 1

        today = date.today()
        for days_ago, score in enumerate(
            [
                Decimal("82.5"),
                Decimal("85.0"),
                Decimal("88.2"),
                Decimal("90.5"),
                Decimal("92.0"),
            ],
            start=1,
        ):
            score_date = today - timedelta(days=days_ago)
            db.add(
                BrigadeDailyScore(
                    brigade_id=brigade.id,
                    score_date=score_date,
                    score=score,
                    details={
                        "productivity": max(0, float(score) - 70),
                        "incidents": 0 if score > Decimal("85.0") else 1,
                    },
                )
            )
            scores_created += 1

    if brigades_created or scores_created:
        await db.commit()

    return brigade_map, brigades_created, scores_created


async def _create_templates(
    db: AsyncSession,
    *,
    current_user: User,
) -> Tuple[Dict[str, ChecklistTemplate], int]:
    """Create demo checklist templates and return them with count."""
    template_map: Dict[str, ChecklistTemplate] = {}
    templates_created = 0

    for payload in DEMO_TEMPLATES:
        result = await db.execute(
            select(ChecklistTemplate).where(ChecklistTemplate.name == payload["name"])
        )
        template = result.scalar_one_or_none()
        if template:
            template_map[payload["name"]] = template
            continue

        template = ChecklistTemplate(
            name=payload["name"],
            description=payload["description"],
            schema=payload["schema"],
            status=TemplateStatus.ACTIVE,
        )
        template.created_by = current_user.id

        db.add(template)
        await db.flush()

        version = ChecklistTemplateVersion(
            template_id=template.id,
            version=template.version,
            schema=payload["schema"],
            diff=None,
            created_by=current_user.id,
        )
        db.add(version)

        template_map[payload["name"]] = template
        templates_created += 1

    if templates_created:
        await db.commit()

    return template_map, templates_created


async def _create_checks_and_reports(
    db: AsyncSession,
    *,
    template_map: Dict[str, ChecklistTemplate],
    brigade_map: Dict[str, Brigade],
    user_map: Dict[str, User],
    current_user: User,
) -> Dict[str, int]:
    """Create demo checks and reports if they are missing."""
    now = datetime.utcnow()
    created_checks = 0
    created_reports = 0

    for payload in DEMO_CHECKS:
        template = template_map.get(payload["template"])
        brigade = brigade_map.get(payload["brigade"])
        inspector = user_map.get(payload["inspector"])
        if not template or not brigade or not inspector:
            continue

        existing_stmt = (
            select(CheckInstance)
            .where(CheckInstance.project_id == payload["project_id"])
            .options(selectinload(CheckInstance.reports))
        )
        result = await db.execute(existing_stmt)
        existing_check = result.scalar_one_or_none()
        if existing_check:
            for report_obj in existing_check.reports:
                try:
                    await _ensure_report_file(report_obj, existing_check)
                except Exception as exc:
                    print(f"[demo] Failed to backfill report {report_obj.id}: {exc}")
            continue

        planned_hours_ahead = payload.get("planned_hours_ahead")
        started_offset = payload.get("started_offset")
        duration_hours = payload.get("duration_hours")

        scheduled_at = None
        started_at = None
        finished_at = None

        if planned_hours_ahead is not None:
            scheduled_at = now + timedelta(hours=planned_hours_ahead)
        elif started_offset is not None:
            started_at = now - timedelta(**started_offset)
            finished_at = (
                started_at + timedelta(hours=duration_hours)
                if duration_hours is not None
                else None
            )
            scheduled_at = started_at - timedelta(hours=12)
        else:
            scheduled_at = now
            started_at = now

        check = CheckInstance(
            template_id=template.id,
            template_version=template.version,
            project_id=payload["project_id"],
            department_id="QA",
            scheduled_at=scheduled_at,
            started_at=started_at,
            finished_at=finished_at,
            inspector_id=inspector.id,
            brigade_id=brigade.id,
            status=payload["status"],
            answers=payload.get("answers", {}),
            comments={"summary": "Generated for demo purposes"},
        )
        db.add(check)
        await db.flush()
        created_checks += 1

        for i, fmt in enumerate(payload["report_formats"]):
            report_format = ReportFormatXLSX(fmt) if isinstance(fmt, str) else fmt
            # Vary report statuses for demo
            if i == 0 and len(payload["report_formats"]) > 1:
                report_status = ReportStatus.GENERATING
            elif i == len(payload["report_formats"]) - 1 and created_reports % 3 == 0:
                report_status = ReportStatus.FAILED
            else:
                report_status = ReportStatus.READY
            
            report = Report(
                check_instance_id=check.id,
                format=report_format,
                file_key=f"demo/{check.id}/{report_format.value}",
                status=report_status,
                generated_by=current_user.id,
                author_id=current_user.id,
                metadata={"source": "demo_seed"},
            )
            db.add(report)
            await db.flush()
            created_reports += 1
            if report_status == ReportStatus.READY:
                try:
                    await _ensure_report_file(report, check)
                except Exception as exc:
                    # Log silently; demo data generation should not fail due to missing storage
                    print(f"[demo] Failed to upload placeholder report {report.id}: {exc}")

    if created_checks or created_reports:
        await db.commit()

    # Get all created check instances for remarks
    check_instances = []
    if created_checks:
        result = await db.execute(
            select(CheckInstance)
            .where(CheckInstance.project_id.in_([p["project_id"] for p in DEMO_CHECKS]))
            .options(selectinload(CheckInstance.reports))
        )
        check_instances = result.scalars().all()

    return {
        "checks_created": created_checks,
        "reports_created": created_reports,
        "check_instances": check_instances,
    }


async def _ensure_report_file(report_obj: Report, check: CheckInstance) -> None:
    """Upload a placeholder file for demo reports if nothing exists yet."""
    try:
        exists = await asyncio.to_thread(storage_service.file_exists, report_obj.file_key)
        if exists:
            return
    except Exception as exc:
        # If storage check fails, assume file doesn't exist and continue
        print(f"[demo] Failed to check if report file exists {report_obj.file_key}: {exc}")

    content_type = "application/octet-stream"
    buffer = BytesIO()

    try:
        if report_obj.format == ReportFormatXLSX.XLSX:
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            wb = Workbook()
            ws = wb.active
            ws.title = "Demo Summary"
            ws["A1"] = "Демо отчет"
            ws["B1"] = str(report_obj.id)
            ws["A2"] = "Проект"
            ws["B2"] = check.project_id or "—"
            ws["A3"] = "Бригада"
            ws["B3"] = str(check.brigade_id) if check.brigade_id else "—"
            ws["A4"] = "Статус обхода"
            ws["B4"] = check.status.value if isinstance(check.status, CheckStatus) else str(check.status)
            ws["A5"] = "Сгенерирован"
            ws["B5"] = datetime.utcnow().isoformat()
            ws["A7"] = "Ответы"

            row = 8
            for key, value in (check.answers or {}).items():
                ws[f"A{row}"] = key
                ws[f"B{row}"] = str(value)
                row += 1

            wb.save(buffer)
        else:
            buffer.write(b"Demo report placeholder")

        buffer.seek(0)
        await asyncio.to_thread(
            storage_service.upload_fileobj,
            buffer,
            report_obj.file_key,
            content_type,
        )
    except Exception as exc:
        # Log but don't fail - demo data generation should continue even if storage fails
        print(f"[demo] Failed to upload report file {report_obj.file_key}: {exc}")
        # Don't raise - allow demo data generation to continue without storage files

async def _create_schedules(
    db: AsyncSession,
    *,
    template_map: Dict[str, ChecklistTemplate],
    user_map: Dict[str, User],
    brigade_map: Dict[str, Brigade],
) -> None:
    """Create demo schedules for automated check creation."""
    
    safety_template = template_map.get("Demo Template: Safety Walk")
    if not safety_template:
        return
    
    inspector_users = [u for u in user_map.values() if "inspector" in u.email]
    brigade_list = list(brigade_map.values())
    
    if not inspector_users or not brigade_list:
        return
    
    # Check if schedules already exist
    result = await db.execute(select(Schedule).where(Schedule.name.like("Demo Schedule%")))
    existing = result.scalars().first()
    if existing:
        return
    
    final_template = template_map.get("Demo Template: Final Inspection")
    
    schedules = [
        Schedule(
            name="Demo Schedule: Daily Safety Walk",
            template_id=safety_template.id,
            cron_or_rrule="0 9 * * *",  # Every day at 9 AM
            inspector_pool=[u.id for u in inspector_users[:2]],
            brigade_pool=[b.id for b in brigade_list[:2]],
            enabled=True,
            timezone="UTC",
        ),
    ]
    
    if final_template:
        schedules.append(
            Schedule(
                name="Demo Schedule: Weekly Final Inspection",
                template_id=final_template.id,
                cron_or_rrule="0 14 * * 5",  # Every Friday at 2 PM
                inspector_pool=[inspector_users[0].id] if inspector_users else [],
                brigade_pool=[b.id for b in brigade_list],
                enabled=True,
                timezone="UTC",
            )
        )
    
    for schedule in schedules:
        if schedule.template_id:
            db.add(schedule)
    
    await db.commit()


async def _create_webhooks(db: AsyncSession) -> None:
    """Create demo webhook subscriptions."""
    result = await db.execute(select(WebhookSubscription).where(WebhookSubscription.url.like("%demo%")))
    existing = result.scalars().first()
    if existing:
        return
    
    webhooks = [
        WebhookSubscription(
            event=WebhookEvent.CHECK_CREATED,
            url="https://demo.example.com/webhooks/check-created",
            secret="demo-secret-key-123",
            active=True,
        ),
        WebhookSubscription(
            event=WebhookEvent.CHECK_COMPLETED,
            url="https://demo.example.com/webhooks/check-completed",
            secret="demo-secret-key-456",
            active=True,
        ),
        WebhookSubscription(
            event=WebhookEvent.REPORT_READY,
            url="https://demo.example.com/webhooks/report-ready",
            secret="demo-secret-key-789",
            active=True,
        ),
    ]
    
    for webhook in webhooks:
        db.add(webhook)
    
    await db.commit()


async def _create_period_summaries(
    db: AsyncSession,
    *,
    brigade_map: Dict[str, Brigade],
    user_map: Dict[str, User],
) -> None:
    """Create demo period summaries for analytics."""
    today = date.today()
    admin_user = user_map.get("demo.manager@example.com") or list(user_map.values())[0] if user_map else None
    
    if not admin_user:
        return
    
    # Check if summaries already exist
    result = await db.execute(
        select(ReportPeriodSummary).where(ReportPeriodSummary.author_id == admin_user.id)
    )
    existing = result.scalars().first()
    if existing:
        return
    
    summaries = []
    brigade_list = list(brigade_map.values())
    
    # Daily summaries for last 7 days
    for days_ago in range(1, 8):
        summary_date = today - timedelta(days=days_ago)
        if brigade_list:
            summaries.append(
                ReportPeriodSummary(
                    granularity=PeriodSummaryGranularity.DAY,
                    period_start=summary_date,
                    period_end=summary_date,
                    brigade_id=brigade_list[0].id if brigade_list else None,
                    author_id=admin_user.id,
                    report_count=2 + days_ago % 3,
                    summary_metrics={
                        "avg_score": float(85 + days_ago % 10),
                        "total_checks": 2 + days_ago % 3,
                        "completed_checks": 1 + days_ago % 2,
                    },
                    delta_metrics={
                        "score_delta": float((days_ago % 5) - 2),
                        "check_count_delta": (days_ago % 3) - 1,
                    } if days_ago > 1 else None,
                )
            )
    
    # Weekly summary for last week
    week_start = today - timedelta(days=today.weekday() + 7)
    week_end = week_start + timedelta(days=6)
    if brigade_list:
        summaries.append(
            ReportPeriodSummary(
                granularity=PeriodSummaryGranularity.WEEK,
                period_start=week_start,
                period_end=week_end,
                brigade_id=brigade_list[0].id if brigade_list else None,
                author_id=admin_user.id,
                report_count=15,
                summary_metrics={
                    "avg_score": 87.5,
                    "total_checks": 15,
                    "completed_checks": 12,
                },
                delta_metrics={
                    "score_delta": 2.3,
                    "check_count_delta": 3,
                },
            )
        )
    
    # Monthly summary for last month
    month_start = date(today.year, today.month, 1) - timedelta(days=30)
    month_end = date(today.year, today.month, 1) - timedelta(days=1)
    if brigade_list:
        summaries.append(
            ReportPeriodSummary(
                granularity=PeriodSummaryGranularity.MONTH,
                period_start=month_start,
                period_end=month_end,
                brigade_id=brigade_list[0].id if brigade_list else None,
                author_id=admin_user.id,
                report_count=45,
                summary_metrics={
                    "avg_score": 86.2,
                    "total_checks": 45,
                    "completed_checks": 38,
                },
                delta_metrics={
                    "score_delta": 1.8,
                    "check_count_delta": 5,
                },
            )
        )
    
    for summary in summaries:
        db.add(summary)
    
    await db.commit()


async def _create_report_generation_events(db: AsyncSession) -> None:
    """Create demo report generation events."""
    result = await db.execute(
        select(ReportGenerationEvent).where(ReportGenerationEvent.event_type == ReportGenerationEventType.MANUAL)
    )
    existing = result.scalars().first()
    if existing:
        return
    
    # Get some reports to attach events to
    reports_result = await db.execute(select(Report).limit(5))
    reports = reports_result.scalars().all()
    
    if not reports:
        return
    
    events = []
    for i, report in enumerate(reports[:3]):
        events.append(
            ReportGenerationEvent(
                report_id=report.id,
                check_instance_id=report.check_instance_id,
                event_type=ReportGenerationEventType.MANUAL if i == 0 else ReportGenerationEventType.SCHEDULED,
                status=ReportGenerationStatus.SUCCESS,
                triggered_by="demo_user",
                payload={"demo": True},
                completed_at=datetime.utcnow() - timedelta(hours=i),
            )
        )
    
    # Add one failed event
    if len(reports) > 3:
        events.append(
            ReportGenerationEvent(
                report_id=reports[3].id if len(reports) > 3 else reports[0].id,
                check_instance_id=reports[3].check_instance_id if len(reports) > 3 else reports[0].check_instance_id,
                event_type=ReportGenerationEventType.RETRY,
                status=ReportGenerationStatus.FAILED,
                triggered_by="demo_user",
                payload={"demo": True, "retry_count": 1},
                error_message="Demo error: Storage service temporarily unavailable",
            )
        )
    
    for event in events:
        db.add(event)
    
    await db.commit()


async def _create_remarks(db: AsyncSession, check_instances: List[CheckInstance]) -> None:
    """Create demo remarks for check instances."""
    if not check_instances:
        return
    
    result = await db.execute(
        select(RemarkEntry).where(RemarkEntry.source == "demo")
    )
    existing = result.scalars().first()
    if existing:
        return
    
    remarks = []
    for i, check in enumerate(check_instances[:5]):
        if check.status == CheckStatus.COMPLETED:
            severity = RemarkSeverity.MEDIUM if i % 2 == 0 else RemarkSeverity.LOW
            if i == 0:
                severity = RemarkSeverity.HIGH
            
            remarks.append(
                RemarkEntry(
                    check_instance_id=check.id,
                    department_id=check.department_id,
                    brigade_id=check.brigade_id,
                    severity=severity,
                    message=f"Демо-замечание #{i+1}: {'Требуется внимание' if severity == RemarkSeverity.HIGH else 'Рекомендация по улучшению'}",
                    raised_at=check.finished_at or check.started_at or datetime.utcnow(),
                    source="demo",
                    details={"demo": True, "check_project": check.project_id},
                )
            )
    
    for remark in remarks:
        db.add(remark)
    
    if remarks:
        await db.commit()


async def generate_demo_data(db: AsyncSession, current_user: User, locale: str = "en") -> Dict[str, Any]:
    """Generate or reuse demo entities for showcasing the system."""
    counters = DemoDataResult()

    required_roles = {"inspector", "crew_leader", "viewer", "admin"}
    role_map = await ensure_roles(db, role_names=required_roles)

    user_map, counters.users_created = await _get_or_create_users(
        db, role_map=role_map
    )

    brigade_map, counters.brigades_created, counters.scores_created = (
        await _create_brigades(db, user_map=user_map)
    )

    template_map, counters.templates_created = await _create_templates(
        db, current_user=current_user
    )

    checks_reports = await _create_checks_and_reports(
        db,
        template_map=template_map,
        brigade_map=brigade_map,
        user_map=user_map,
        current_user=current_user,
    )
    counters.checks_created = checks_reports["checks_created"]
    counters.reports_created = checks_reports["reports_created"]

    # Create additional demo features
    await _create_schedules(db, template_map=template_map, user_map=user_map, brigade_map=brigade_map)
    await _create_webhooks(db)
    await _create_period_summaries(db, brigade_map=brigade_map, user_map=user_map)
    await _create_report_generation_events(db)
    await _create_remarks(db, checks_reports.get("check_instances", []))

    return counters.as_payload(locale=locale)


async def reset_project_to_clean_state(db: AsyncSession) -> Dict[str, Any]:
    """Wipe project data and recreate only the default administrator account."""
    from app.services.reset_service import reset_service

    results = await reset_service.reset_project(db)

    # Verify reset
    verification = await reset_service.verify_reset(db)

    return {
        "status": "reset",
        "detail": "Создан чистый проект с тестовым администратором.",
        "tables_truncated": len(results.get("tables_truncated", [])),
        "tables_preserved": len(results.get("tables_preserved", [])),
        "admin_user_preserved": results.get("admin_user_preserved", False),
        "admin_email": results.get("admin_email", DEFAULT_ADMIN_EMAIL),
        "admin_password": DEFAULT_ADMIN_PASSWORD,
        "verification": verification,
        "errors": results.get("errors", []),
    }


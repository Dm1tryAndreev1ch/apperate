"""Report generation service."""
from __future__ import annotations

import io
from typing import Any, Dict, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.models.report import ReportFormatXLSX
from app.models.checklist import CheckInstance
from app.services.storage_service import storage_service


class ReportService:
    """Service for generating Excel reports (XLSX-only)."""

    SUMMARY_SHEET = "Summary"
    ANSWERS_SHEET = "Answers"

    @staticmethod
    def _auto_size_columns(sheet) -> None:
        """Auto-fit column widths based on content length."""
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            adjusted_width = min(max_length + 4, 60)
            sheet.column_dimensions[column].width = adjusted_width

    @staticmethod
    def _populate_summary_sheet(
        sheet,
        *,
        check_instance: CheckInstance,
        inspector_name: str,
        template_schema: Dict[str, Any],
    ) -> None:
        """Fill the summary sheet with metadata."""
        sheet["A1"].value = "Отчет MantaQC"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet.merge_cells("A1:D1")
        sheet["A1"].alignment = Alignment(horizontal="left")

        summary_rows = [
            ("ID обхода", str(check_instance.id)),
            ("Шаблон", str(template_schema.get("name") or check_instance.template_id)),
            ("Версия шаблона", check_instance.template_version),
            ("Инспектор", inspector_name or "—"),
            ("Статус", getattr(check_instance.status, "value", str(check_instance.status))),
            ("Начало", check_instance.started_at.isoformat() if check_instance.started_at else "—"),
            ("Окончание", check_instance.finished_at.isoformat() if check_instance.finished_at else "—"),
            ("Проект", check_instance.project_id or "—"),
            ("Подразделение", check_instance.department_id or "—"),
        ]

        for idx, (label, value) in enumerate(summary_rows, start=3):
            sheet[f"A{idx}"].value = label
            sheet[f"A{idx}"].font = Font(bold=True)
            sheet[f"B{idx}"].value = value

        ReportService._auto_size_columns(sheet)

    @staticmethod
    def _populate_answers_sheet(sheet, *, check_instance: CheckInstance, template_schema: Dict[str, Any]) -> None:
        """Fill the answers sheet with question breakdown."""
        sheet["A1"].value = "Вопрос"
        sheet["B1"].value = "Ответ"
        sheet["C1"].value = "Комментарий"
        sheet["A1"].font = sheet["B1"].font = sheet["C1"].font = Font(bold=True)

        answers = check_instance.answers or {}
        comments = check_instance.comments or {}

        row = 2
        sections: Iterable[Dict[str, Any]] = template_schema.get("sections", [])
        for section in sections:
            for question in section.get("questions", []):
                question_id = question.get("id")
                sheet[f"A{row}"].value = question.get("text") or question_id
                sheet[f"B{row}"].value = answers.get(question_id, "—")
                sheet[f"C{row}"].value = comments.get(question_id, comments.get("summary"))
                row += 1

        ReportService._auto_size_columns(sheet)

    @staticmethod
    def generate_xlsx(
        check_instance: CheckInstance,
        template_schema: Dict[str, Any],
        inspector_name: str = "Unknown",
    ) -> bytes:
        """Generate a single-sheet XLSX report."""
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = ReportService.SUMMARY_SHEET
        ReportService._populate_summary_sheet(
            summary_sheet,
            check_instance=check_instance,
            inspector_name=inspector_name,
            template_schema=template_schema,
        )

        answers_sheet = workbook.create_sheet(ReportService.ANSWERS_SHEET)
        ReportService._populate_answers_sheet(
            answers_sheet,
            check_instance=check_instance,
            template_schema=template_schema,
        )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    @staticmethod
    def generate_and_upload(
        check_instance: CheckInstance,
        template_schema: Dict[str, Any],
        format: ReportFormatXLSX,
        inspector_name: str = "Unknown",
    ) -> str:
        """Generate report and upload to S3, return S3 key."""
        if format != ReportFormatXLSX.XLSX:
            raise ValueError(f"Unsupported format: {format}")

        payload = ReportService.generate_xlsx(check_instance, template_schema, inspector_name)
        file_obj = io.BytesIO(payload)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # Generate S3 key
        file_extension = format.value
        key = f"reports/{check_instance.id}/{format.value}/{check_instance.id}.{file_extension}"

        file_obj.seek(0)
        storage_service.upload_fileobj(file_obj, key, content_type=content_type)
        return key


report_service = ReportService()


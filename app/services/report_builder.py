"""Report builder service for generating Excel reports with charts and analytics."""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.checklist import CheckInstance
from app.services.analytics_service import (
    AlertDTO,
    BrigadeScoreDTO,
    PeriodSummaryDTO,
    ReportAnalyticsDTO,
)


class ReportBuilder:
    """Service for building Excel reports with charts and formatting."""

    # Color scheme matching MantaQC branding
    HEADER_FILL = PatternFill(start_color="173F5F", end_color="173F5F", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    TITLE_FONT = Font(bold=True, size=14)
    SUBTITLE_FONT = Font(bold=True, size=12)

    @staticmethod
    def build_report_workbook(
        *,
        check_instance: CheckInstance,
        analytics: ReportAnalyticsDTO,
        inspector_name: str = "Unknown",
        template_name: str = "Template",
    ) -> bytes:
        """Build a complete Excel workbook for a check instance report."""
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Cover sheet
        cover_sheet = workbook.create_sheet("MantaQC — Сводный отчёт")
        ReportBuilder._populate_cover_sheet(
            cover_sheet,
            check_instance=check_instance,
            analytics=analytics,
            inspector_name=inspector_name,
            template_name=template_name,
        )

        # Analytics sheet
        analytics_sheet = workbook.create_sheet("Аналитика")
        ReportBuilder._populate_analytics_sheet(
            analytics_sheet,
            analytics=analytics,
        )

        # Issues sheet
        if analytics.alerts:
            issues_sheet = workbook.create_sheet("Проблемы")
            ReportBuilder._populate_issues_sheet(
                issues_sheet,
                alerts=analytics.alerts,
                check_instance=check_instance,
            )

        # Checks sheet (detailed checklist data)
        checks_sheet = workbook.create_sheet("Обходы")
        ReportBuilder._populate_checks_sheet(
            checks_sheet,
            check_instance=check_instance,
        )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    @staticmethod
    def _populate_cover_sheet(
        sheet,
        *,
        check_instance: CheckInstance,
        analytics: ReportAnalyticsDTO,
        inspector_name: str,
        template_name: str,
    ) -> None:
        """Populate the cover sheet with metadata and KPIs."""
        # Title
        sheet["A1"].value = "MantaQC — Сводный отчёт"
        sheet["A1"].font = ReportBuilder.TITLE_FONT
        sheet.merge_cells("A1:D1")

        # Metadata section
        row = 3
        metadata = [
            ("ID обхода", str(check_instance.id)),
            ("Шаблон", template_name),
            ("Версия шаблона", check_instance.template_version),
            ("Инспектор", inspector_name),
            ("Статус", getattr(check_instance.status, "value", str(check_instance.status))),
            ("Начало", check_instance.started_at.isoformat() if check_instance.started_at else "—"),
            ("Окончание", check_instance.finished_at.isoformat() if check_instance.finished_at else "—"),
            ("Проект", check_instance.project_id or "—"),
            ("Подразделение", check_instance.department_id or "—"),
        ]

        for label, value in metadata:
            sheet[f"A{row}"].value = label
            sheet[f"A{row}"].font = Font(bold=True)
            sheet[f"B{row}"].value = value
            row += 1

        # KPI cards section
        row += 1
        sheet[f"A{row}"].value = "Основные показатели"
        sheet[f"A{row}"].font = ReportBuilder.SUBTITLE_FONT
        row += 1

        kpi_data = [
            ("Средний балл", analytics.avg_score),
            ("Балл бригады", analytics.brigade_score.score if analytics.brigade_score else None),
            ("Замечания", analytics.remark_count),
            ("Критические нарушения", len(analytics.critical_violations)),
        ]

        for idx, (label, value) in enumerate(kpi_data):
            col = "A" if idx % 2 == 0 else "C"
            if idx >= 2:
                col_offset = 0 if idx % 2 == 0 else 2
                row_offset = 2 if idx >= 2 else 0
                actual_row = row + row_offset
            else:
                actual_row = row

            sheet[f"{col}{actual_row}"].value = label
            sheet[f"{col}{actual_row}"].font = Font(bold=True)
            sheet[f"{get_column_letter(ord(col) + 1)}{actual_row}"].value = (
                float(value) if value is not None else "—"
            )

        ReportBuilder._auto_size_columns(sheet)

    @staticmethod
    def _populate_analytics_sheet(
        sheet,
        *,
        analytics: ReportAnalyticsDTO,
    ) -> None:
        """Populate analytics sheet with brigade scores and charts."""
        # Title
        sheet["A1"].value = "Аналитика по культуре производства"
        sheet["A1"].font = ReportBuilder.TITLE_FONT
        sheet.merge_cells("A1:D1")

        # Headers
        headers = ["Бригада", "Дата", "Балл", "Общий балл"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = header
            cell.fill = ReportBuilder.HEADER_FILL
            cell.font = ReportBuilder.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row = 4
        if analytics.brigade_score:
            sheet.cell(row=row, column=1).value = analytics.brigade_score.brigade_name
            sheet.cell(row=row, column=2).value = analytics.brigade_score.score_date.isoformat()
            sheet.cell(row=row, column=3).value = float(analytics.brigade_score.score)
            sheet.cell(row=row, column=4).value = (
                float(analytics.brigade_score.overall_score) if analytics.brigade_score.overall_score else None
            )
            row += 1

        # Add chart
        if analytics.brigade_score:
            chart = LineChart()
            chart.title = "Динамика балла бригады"
            chart.style = 10
            chart.y_axis.title = "Балл"
            chart.x_axis.title = "Дата"

            data = Reference(sheet, min_col=3, min_row=3, max_row=row - 1)
            cats = Reference(sheet, min_col=2, min_row=4, max_row=row - 1)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)

            sheet.add_chart(chart, "F3")

        ReportBuilder._auto_size_columns(sheet)

    @staticmethod
    def _populate_issues_sheet(
        sheet,
        *,
        alerts: List[AlertDTO],
        check_instance: CheckInstance,
    ) -> None:
        """Populate issues sheet with alerts and Bitrix status."""
        # Title
        sheet["A1"].value = "Выявленные проблемы"
        sheet["A1"].font = ReportBuilder.TITLE_FONT
        sheet.merge_cells("A1:F1")

        # Headers
        headers = ["Серьёзность", "Категория", "Сообщение", "ID обхода", "Бригада", "Статус Bitrix"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = header
            cell.fill = ReportBuilder.HEADER_FILL
            cell.font = ReportBuilder.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row = 4
        for alert in alerts:
            sheet.cell(row=row, column=1).value = alert.severity
            sheet.cell(row=row, column=2).value = alert.category
            sheet.cell(row=row, column=3).value = alert.message
            sheet.cell(row=row, column=4).value = str(alert.check_instance_id) if alert.check_instance_id else "—"
            sheet.cell(row=row, column=5).value = str(alert.brigade_id) if alert.brigade_id else "—"
            sheet.cell(row=row, column=6).value = "Создана" if alert.metadata and "bitrix_ticket_id" in alert.metadata else "Ожидает"
            row += 1

        ReportBuilder._auto_size_columns(sheet)

    @staticmethod
    def _populate_checks_sheet(
        sheet,
        *,
        check_instance: CheckInstance,
    ) -> None:
        """Populate checks sheet with detailed checklist data."""
        # Title
        sheet["A1"].value = "Детали обхода"
        sheet["A1"].font = ReportBuilder.TITLE_FONT
        sheet.merge_cells("A1:E1")

        # Headers
        headers = ["Секция", "Вопрос", "Ответ", "Комментарий", "Фото"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = header
            cell.fill = ReportBuilder.HEADER_FILL
            cell.font = ReportBuilder.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row = 4
        template_schema = check_instance.template.schema if check_instance.template else {}
        answers = check_instance.answers or {}
        comments = check_instance.comments or {}
        media_keys = check_instance.media_keys or []

        for section in template_schema.get("sections", []):
            section_name = section.get("title") or section.get("name", "Без названия")
            for question in section.get("questions", []):
                question_id = question.get("id")
                question_text = question.get("text") or question_id
                answer = answers.get(question_id, "—")
                comment = comments.get(question_id) or comments.get("summary") or "—"
                has_media = "yes" if question_id in [str(k) for k in media_keys] else "—"

                sheet.cell(row=row, column=1).value = section_name
                sheet.cell(row=row, column=2).value = question_text
                sheet.cell(row=row, column=3).value = str(answer) if answer is not None else "—"
                sheet.cell(row=row, column=4).value = str(comment) if comment else "—"
                sheet.cell(row=row, column=5).value = has_media
                row += 1

        ReportBuilder._auto_size_columns(sheet)

    @staticmethod
    def build_period_summary_workbook(
        *,
        summary: PeriodSummaryDTO,
    ) -> bytes:
        """Build Excel workbook for period summary."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Сводка за период"

        # Title
        sheet["A1"].value = f"MantaQC — Сводка за {summary.granularity}"
        sheet["A1"].font = ReportBuilder.TITLE_FONT
        sheet.merge_cells("A1:D1")

        # Period info
        row = 3
        sheet[f"A{row}"].value = "Период"
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"B{row}"].value = f"{summary.period_start.isoformat()} — {summary.period_end.isoformat()}"

        row += 1
        sheet[f"A{row}"].value = "Количество отчётов"
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"B{row}"].value = summary.report_count

        row += 1
        sheet[f"A{row}"].value = "Средний балл"
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"B{row}"].value = float(summary.avg_score) if summary.avg_score else "—"

        row += 1
        sheet[f"A{row}"].value = "Количество замечаний"
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"B{row}"].value = summary.remark_count

        # Brigade scores table
        if summary.brigade_scores:
            row += 2
            sheet[f"A{row}"].value = "Баллы бригад"
            sheet[f"A{row}"].font = ReportBuilder.SUBTITLE_FONT

            row += 1
            headers = ["Бригада", "Дата", "Балл", "Общий балл"]
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.value = header
                cell.fill = ReportBuilder.HEADER_FILL
                cell.font = ReportBuilder.HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

            row += 1
            for brigade_score in summary.brigade_scores:
                sheet.cell(row=row, column=1).value = brigade_score.brigade_name
                sheet.cell(row=row, column=2).value = brigade_score.score_date.isoformat()
                sheet.cell(row=row, column=3).value = float(brigade_score.score)
                sheet.cell(row=row, column=4).value = (
                    float(brigade_score.overall_score) if brigade_score.overall_score else None
                )
                row += 1

        # Delta metrics
        if summary.delta_metrics:
            row += 2
            sheet[f"A{row}"].value = "Изменения"
            sheet[f"A{row}"].font = ReportBuilder.SUBTITLE_FONT
            row += 1

            for metric_name, delta_value in summary.delta_metrics.items():
                sheet.cell(row=row, column=1).value = metric_name
                sheet.cell(row=row, column=1).font = Font(bold=True)
                sheet.cell(row=row, column=2).value = float(delta_value)
                row += 1

        ReportBuilder._auto_size_columns(sheet)

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    @staticmethod
    def _auto_size_columns(sheet) -> None:
        """Auto-fit column widths based on content."""
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            adjusted_width = min(max_length + 4, 60)
            sheet.column_dimensions[column].width = adjusted_width


report_builder = ReportBuilder()


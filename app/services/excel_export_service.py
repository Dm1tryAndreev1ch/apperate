"""Services for exporting analytics to Excel."""
from __future__ import annotations

import asyncio
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Dict, Iterable, List, Optional
from uuid import uuid4

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import Select, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.brigade import Brigade, BrigadeDailyScore
from app.services.storage_service import storage_service

NUMBER_FORMAT = "0.0"


@dataclass
class MonthlyBrigadeMetrics:
    """Container with brigade score metrics for a month."""

    brigade_name: str
    daily_scores: Dict[int, Optional[Decimal]]
    current_avg: Optional[Decimal]
    previous_avg: Optional[Decimal]

    @property
    def delta(self) -> Optional[Decimal]:
        if self.current_avg is None or self.previous_avg is None:
            return None
        return self.current_avg - self.previous_avg


async def _fetch_brigades(db: AsyncSession) -> List[Brigade]:
    result = await db.execute(select(Brigade).order_by(Brigade.name))
    return list(result.scalars().all())


async def _fetch_scores(
    db: AsyncSession,
    *,
    brigade_ids: Iterable,
    start_date: date,
    end_date: date,
) -> Dict[str, Dict[date, Decimal]]:
    if not brigade_ids:
        return {}

    stmt: Select = (
        select(
            BrigadeDailyScore.brigade_id,
            BrigadeDailyScore.score_date,
            BrigadeDailyScore.score,
        )
        .where(
            BrigadeDailyScore.brigade_id.in_(brigade_ids),
            BrigadeDailyScore.score_date >= start_date,
            BrigadeDailyScore.score_date <= end_date,
        )
    )
    rows = await db.execute(stmt)
    data: Dict[str, Dict[date, Decimal]] = {}
    for brigade_id, score_date, score in rows:
        data.setdefault(str(brigade_id), {})[score_date] = Decimal(score) if score is not None else None
    return data


async def _fetch_month_average(
    db: AsyncSession,
    *,
    brigade_ids: Iterable,
    start_date: date,
    end_date: date,
) -> Dict[str, Optional[Decimal]]:
    if not brigade_ids:
        return {}

    stmt: Select = (
        select(
            BrigadeDailyScore.brigade_id,
            func.avg(BrigadeDailyScore.score),
        )
        .where(
            BrigadeDailyScore.brigade_id.in_(brigade_ids),
            BrigadeDailyScore.score_date >= start_date,
            BrigadeDailyScore.score_date <= end_date,
        )
        .group_by(BrigadeDailyScore.brigade_id)
    )
    rows = await db.execute(stmt)
    return {
        str(brigade_id): (Decimal(avg_score) if avg_score is not None else None)
        for brigade_id, avg_score in rows
    }


async def _collect_monthly_metrics(
    db: AsyncSession,
    *,
    month: date,
    brigade_filter: Optional[Iterable[str]] = None,
) -> List[MonthlyBrigadeMetrics]:
    month_start = month.replace(day=1)
    _, days_in_month = monthrange(month_start.year, month_start.month)
    month_end = month_start + timedelta(days=days_in_month - 1)

    prev_month_start = (month_start - relativedelta(months=1)).replace(day=1)
    _, prev_days = monthrange(prev_month_start.year, prev_month_start.month)
    prev_month_end = prev_month_start + timedelta(days=prev_days - 1)

    brigades = await _fetch_brigades(db)
    if brigade_filter:
        selected = {str(b_id) for b_id in brigade_filter}
        brigades = [brig for brig in brigades if str(brig.id) in selected]

    brigade_ids = [brig.id for brig in brigades]
    daily_scores = await _fetch_scores(
        db,
        brigade_ids=brigade_ids,
        start_date=month_start,
        end_date=month_end,
    )
    current_avg = await _fetch_month_average(
        db,
        brigade_ids=brigade_ids,
        start_date=month_start,
        end_date=month_end,
    )
    previous_avg = await _fetch_month_average(
        db,
        brigade_ids=brigade_ids,
        start_date=prev_month_start,
        end_date=prev_month_end,
    )

    metrics: List[MonthlyBrigadeMetrics] = []
    for brigade in brigades:
        daily_map = {
            day: daily_scores.get(str(brigade.id), {}).get(month_start + timedelta(days=day - 1))
            for day in range(1, days_in_month + 1)
        }
        metrics.append(
            MonthlyBrigadeMetrics(
                brigade_name=brigade.name,
                daily_scores=daily_map,
                current_avg=current_avg.get(str(brigade.id)),
                previous_avg=previous_avg.get(str(brigade.id)),
            )
        )
    return metrics


def _build_workbook(
    *,
    metrics: List[MonthlyBrigadeMetrics],
    month: date,
) -> BytesIO:
    month_start = month.replace(day=1)
    _, days_in_month = monthrange(month_start.year, month_start.month)

    wb = Workbook()
    ws = wb.active
    ws.title = "Аналитика"

    header_fill = PatternFill(start_color="173F5F", end_color="173F5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A1"] = "Аналитика по культуре производства за месяц"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_in_month + 4)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers: List[str] = ["Структурное подразделение"]
    headers.extend([str(day) for day in range(1, days_in_month + 1)])
    headers.extend(["Итог месяца", "Предыдущий месяц", "Динамика"])

    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = 15 if col_idx == 1 else 10

    for metric in metrics:
        row = [metric.brigade_name]
        for day in range(1, days_in_month + 1):
            score = metric.daily_scores.get(day)
            row.append(float(score) if score is not None else None)
        row.append(float(metric.current_avg) if metric.current_avg is not None else None)
        row.append(float(metric.previous_avg) if metric.previous_avg is not None else None)
        delta = metric.delta
        row.append(float(delta) if delta is not None else None)
        ws.append(row)

    for row in ws.iter_rows(min_row=3, min_col=2, max_col=days_in_month + 4):
        for cell in row:
            if cell.value is not None:
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="center")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


async def generate_monthly_culture_report(
    db: AsyncSession,
    *,
    month: date,
    brigade_ids: Optional[Iterable[str]] = None,
    expires_in: int = 3600,
) -> Dict[str, str]:
    """Generate Excel report for monthly brigade culture metrics."""
    metrics = await _collect_monthly_metrics(db, month=month, brigade_filter=brigade_ids)
    workbook_io = await asyncio.to_thread(_build_workbook, metrics=metrics, month=month)

    file_key = f"reports/monthly_culture/{month.year}-{month.month:02d}-{uuid4()}.xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    await asyncio.to_thread(
        storage_service.upload_fileobj,
        workbook_io,
        file_key,
        content_type,
    )

    download_url = storage_service.generate_download_url(file_key, expires_in=expires_in)

    filename = f"monthly-culture-{month.year}-{month.month:02d}.xlsx"
    return {
        "file_key": file_key,
        "download_url": download_url,
        "filename": filename,
    }

